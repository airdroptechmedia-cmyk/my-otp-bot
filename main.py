import os
import sqlite3
import time
import requests
import threading
import io
import datetime
import openpyxl
import asyncio
import aiohttp
import re
from flask import Flask
from telebot import TeleBot, types
from contextlib import contextmanager

# ==================== KEEP-ALIVE & 1-MIN SELF-PING ENGINE ====================
app = Flask('')

@app.route('/')
def home():
    return "OTP RECIVER PRO BOT with ASYNC DUAL API is Alive 24/7!"

def run_flask():
    try:
        app.run(host='0.0.0.0', port=8080)
    except Exception as e:
        print(f"Flask Server Error: {e}")

def self_ping():
    time.sleep(10)
    url = "https://my-otp-bot-d7jk.onrender.com"  # আপনার Render ওয়েবসাইট লিংক
    while True:
        try:
            requests.get(url, timeout=10)
            print("🚀 1-Min Self-ping successful! Render kept awake.")
        except Exception:
            pass
        time.sleep(60)

threading.Thread(target=run_flask, daemon=True).start()
threading.Thread(target=self_ping, daemon=True).start()

# ==================== CONFIGURATION ====================
BOT_TOKEN = "8842802759:AAFTzG_yyzHiirBiW2Canl2l0t_sG2HxKt8" # বটের টোকেন
ADMIN_ID = 8125384914                                       # Admin ID
BOT_NAME = "OTP RECIVER PRO BOT"
DB_NAME = "fresh_master_shop.db"
# =======================================================

bot = TeleBot(BOT_TOKEN, parse_mode="HTML")

BOT_USERNAME = ""
try:
    BOT_USERNAME = bot.get_me().username
except Exception:
    BOT_USERNAME = "otpreciverpro_bot"

USER_STATES = {}

def set_user_state(user_id, key, value):
    if user_id not in USER_STATES:
        USER_STATES[user_id] = {}
    USER_STATES[user_id][key] = value

def get_user_state(user_id, key, default=None):
    return USER_STATES.get(user_id, {}).get(key, default)

def clear_user_state(user_id):
    if user_id in USER_STATES:
        USER_STATES[user_id].clear()

@contextmanager
def get_db():
    conn = sqlite3.connect(DB_NAME, timeout=15)
    try:
        yield conn
    finally:
        conn.close()

# ----------------- TELEGRAM CLOUD AUTO-BACKUP & RESTORE ENGINE -----------------
LAST_BACKUP_TIME = 0

def backup_db_to_telegram(force=False):
    global LAST_BACKUP_TIME
    now = time.time()
    if not force and (now - LAST_BACKUP_TIME < 10):
        return
        
    try:
        if os.path.exists(DB_NAME):
            with open(DB_NAME, 'rb') as doc:
                bot.send_document(
                    ADMIN_ID,
                    doc,
                    caption=f"#DB_BACKUP | Auto Cloud Sync | {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                    disable_notification=True
                )
            LAST_BACKUP_TIME = now
            print("☁️ DB Cloud Backup sent to Telegram successfully!")
    except Exception as e:
        print(f"Backup DB Exception: {e}")

def restore_db_from_telegram():
    try:
        updates = bot.get_updates(limit=100)
        for update in reversed(updates):
            msg = update.message
            if msg and msg.caption and "#DB_BACKUP" in msg.caption and msg.document:
                file_info = bot.get_file(msg.document.file_id)
                downloaded = bot.download_file(file_info.file_path)
                with open(DB_NAME, 'wb') as f:
                    f.write(downloaded)
                print("🎉 DATABASE RESTORED SUCCESSFULLY FROM TELEGRAM CLOUD!")
                return True
    except Exception as e:
        print(f"Restore DB Exception: {e}")
    return False

restore_db_from_telegram()

def create_excel_document(product_name, lines):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accounts"
    
    for idx, line in enumerate(lines, start=1):
        ws.cell(row=idx, column=1, value=line)
        
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    clean_name = product_name.replace(" ", "_")
    file_name = f"{clean_name}_{now_str}.xlsx"
    stream.name = file_name
    return stream

# ----------------- DATABASE INITIALIZATION -----------------
def init_db():
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("PRAGMA journal_mode=WAL;")
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            balance REAL DEFAULT 0.00,
            reward_balance REAL DEFAULT 0.00,
            referrals INTEGER DEFAULT 0,
            referred_by INTEGER,
            otp_count INTEGER DEFAULT 0,
            joined_date TEXT DEFAULT CURRENT_DATE
        )''')
        try:
            cursor.execute("ALTER TABLE users ADD COLUMN reward_balance REAL DEFAULT 0.00")
        except Exception:
            pass
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS active_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            order_id TEXT,
            phone_number TEXT,
            service TEXT,
            country TEXT,
            status TEXT DEFAULT 'WAITING',
            last_code TEXT DEFAULT NULL,
            api_source TEXT DEFAULT 'API1',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        try:
            cursor.execute("ALTER TABLE active_orders ADD COLUMN api_source TEXT DEFAULT 'API1'")
        except Exception:
            pass
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS purchases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            product_name TEXT,
            category TEXT,
            qty INTEGER,
            total_price REAL,
            content TEXT,
            date TEXT DEFAULT CURRENT_DATE
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS withdrawals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            amount REAL,
            method TEXT,
            account_number TEXT,
            status TEXT DEFAULT 'pending',
            date TEXT DEFAULT CURRENT_DATE
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS countries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            service_name TEXT,
            service_code TEXT DEFAULT '26134',
            country_name TEXT,
            country_flag TEXT DEFAULT '🌐',
            country_code TEXT,
            price REAL DEFAULT 0.00
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT,
            subcategory TEXT DEFAULT '',
            name TEXT,
            price REAL
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS item_stock (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER,
            content TEXT,
            status TEXT DEFAULT 'available'
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS deposits (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            amount REAL,
            method TEXT,
            trx_id TEXT UNIQUE,
            status TEXT DEFAULT 'pending'
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )''')
        
        defaults = {
            'bkash_num': '01625212609',
            'nagad_num': '01625212609',
            'binance_uid': '1133157464',
            'min_deposit': '20.0',
            'min_withdraw': '50.0',
            'deposit_bonus': '5',
            'refer_reward': '0.11',
            'otp_reward': '0.10',
            'signup_bonus': '0.00',
            
            'number_api_key': 'M7D4REK5Y06',
            'api_base_url': 'https://api.2oo9.cloud/MXS47FLFX0U/tnevs/@public/api',
            'getnum_url': 'https://api.2oo9.cloud/MXS47FLFX0U/tnevs/@public/api/getnum',
            'getmsg_url': 'https://api.2oo9.cloud/MXS47FLFX0U/tnevs/@public/api/success-otp',
            
            'number_api_key_2': 'M6SB7HZXXIX',
            'api_base_url_2': 'https://api.2oo9.cloud/MXS47FLFX0U/tnevs/@public/api',
            'getnum_url_2': 'https://api.2oo9.cloud/MXS47FLFX0U/tnevs/@public/api/getnum',
            'getmsg_url_2': 'https://api.2oo9.cloud/MXS47FLFX0U/tnevs/@public/api/success-otp',
            
            'traffic_url': 'https://api.2oo9.cloud/MXS47FLFX0U/tnevs/@public/api/console',
            'force_channels': '',
            'otp_group_id': '@otpreciverpro',
            'otp_group_link': 'https://t.me/otpreciverpro'
        }
        for k, v in defaults.items():
            cursor.execute("INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)", (k, v))

        cursor.execute("UPDATE settings SET value='M7D4REK5Y06' WHERE key='number_api_key' AND (value='M9NA8XX44CT' OR value='')")
        cursor.execute("UPDATE settings SET value='M6SB7HZXXIX' WHERE key='number_api_key_2' AND value=''")
            
        cursor.execute("SELECT COUNT(*) FROM countries")
        if cursor.fetchone()[0] == 0:
            default_countries = [
                ('Facebook', '26134', 'Uzbekistan', '🇺🇿', 'uz', 0.00),
                ('Facebook', '26134', 'United Kingdom', '🇬🇧', 'uk', 0.00),
                ('Instagram', '26135', 'Uzbekistan', '🇺🇿', 'uz', 0.00),
                ('Telegram', '26136', 'Uzbekistan', '🇺🇿', 'uz', 0.00)
            ]
            cursor.executemany("INSERT INTO countries (service_name, service_code, country_name, country_flag, country_code, price) VALUES (?, ?, ?, ?, ?, ?)", default_countries)

        conn.commit()

init_db()

def get_setting(key):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT value FROM settings WHERE key=?", (key,))
        res = cursor.fetchone()
        return res[0] if res else ""

def set_setting(key, value):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, str(value)))
        conn.commit()
    backup_db_to_telegram()

# ----------------- STRICT FORCE JOIN CHECKER -----------------
def is_user_joined(user_id):
    if user_id == ADMIN_ID:
        return True
    raw_channels = get_setting('force_channels')
    if not raw_channels:
        return True
    
    channels = [c.strip() for c in raw_channels.split(',') if c.strip()]
    for ch in channels:
        try:
            ch_uname = ch if ch.startswith("@") else f"@{ch}"
            member = bot.get_chat_member(ch_uname, user_id)
            if member.status not in ['creator', 'administrator', 'member']:
                return False
        except Exception:
            pass
    return True

def send_force_join_msg(chat_id):
    raw_channels = get_setting('force_channels')
    channels = [c.strip() for c in raw_channels.split(',') if c.strip()]
    
    markup = types.InlineKeyboardMarkup(row_width=1)
    for idx, ch in enumerate(channels, start=1):
        clean_ch = ch.replace("@", "")
        markup.add(types.InlineKeyboardButton(f"📢 Join Channel #{idx}", url=f"https://t.me/{clean_ch}"))
        
    markup.add(types.InlineKeyboardButton("✅ Verify / I Have Joined", callback_data="check_join_verify"))
    bot.send_message(
        chat_id,
        "⚠️ <b>বট ব্যবহার করতে হলে আপনাকে আমাদের অফিশিয়াল চ্যানেলে জয়েন হতে হবে!</b>\n\n"
        "নিচের বাটনে চাপ দিয়ে চ্যানেলে জয়েন করুন, তারপর <b>Verify</b> বাটনে ক্লিক করুন:",
        reply_markup=markup
    )

@bot.callback_query_handler(func=lambda call: call.data == "check_join_verify")
def check_join_verify_cb(call):
    if is_user_joined(call.from_user.id):
        bot.answer_callback_query(call.id, "🎉 ভেরিফিকেশন সফল হয়েছে!", show_alert=True)
        try:
            bot.delete_message(call.message.chat.id, call.message.message_id)
        except Exception:
            pass
        bot.send_message(call.message.chat.id, f"👋 <b>{BOT_NAME}</b>-এ আপনাকে স্বাগতম!", reply_markup=main_reply_keyboard(call.from_user.id))
    else:
        bot.answer_callback_query(call.id, "❌ আপনি এখনো সবগুলো চ্যানেলে জয়েন করেননি! আগে জয়েন করুন।", show_alert=True)

# ----------------- HIGH-SPEED ASYNCHRONOUS OTP EXTRACTION ENGINE -----------------
def extract_otp_from_response(res, clean_phone, order_id):
    if not res:
        return None
    
    clean_phone = str(clean_phone).replace("+", "").strip() if clean_phone else ""
    str_order_id = str(order_id).replace("+", "").strip() if order_id else ""
    
    def extract_code_from_str(s):
        if not s or not isinstance(s, (str, int)):
            return None
        s_str = str(s).strip()
        if not s_str:
            return None
        
        ignored_keywords = [
            "none", "null", "waiting", "pending", "false", "true", "ok", "success", 
            "status_ok", "200", "400", "404", "500", "error", "no_sms", "not_found", 
            "processing", "no_otp", "exist", "number_active", "received", "cancelled",
            "cancel", "timeout", "expire", "expired", "wait", "done"
        ]
        if s_str.lower() in ignored_keywords:
            return None

        if s_str.isdigit() and 4 <= len(s_str) <= 8:
            return s_str

        matches = re.findall(r'\b\d{4,8}\b', s_str)
        if matches:
            return matches[0]

        prefix_match = re.search(r'\b([A-Za-z]{1,4}[- ]?\d{4,8})\b', s_str)
        if prefix_match:
            return prefix_match.group(1)

        hyphen_match = re.search(r'\b\d{3,4}-\d{3,4}\b', s_str)
        if hyphen_match:
            return hyphen_match.group(0).replace('-', '')

        return None

    def get_code_from_dict(d):
        if not isinstance(d, dict):
            return None
        for field in ['otp', 'sms', 'last_code', 'text', 'message', 'msg', 'code']:
            val = d.get(field)
            extracted = extract_code_from_str(val)
            if extracted:
                return extracted
        return None

    def matches_target(d):
        if not isinstance(d, dict):
            return True
        possible_nums = []
        for k in ['full_number', 'number', 'phone', 'no_plus_number', 'order_id', 'id']:
            v = d.get(k)
            if v is not None:
                possible_nums.append(str(v).replace("+", "").strip())
        if not possible_nums:
            return True
        for num in possible_nums:
            if clean_phone and (clean_phone in num or num in clean_phone):
                return True
            if str_order_id and (str_order_id in num or num in str_order_id):
                return True
        return False

    if isinstance(res, (str, int)):
        return extract_code_from_str(res)

    if isinstance(res, dict):
        for key in ["data", "result", "messages", "orders", "sms"]:
            val = res.get(key)
            if val is not None:
                if isinstance(val, list):
                    for item in val:
                        if isinstance(item, dict) and matches_target(item):
                            c = get_code_from_dict(item)
                            if c:
                                return c
                        elif isinstance(item, (str, int)):
                            c = extract_code_from_str(item)
                            if c:
                                return c
                elif isinstance(val, dict):
                    if matches_target(val):
                        c = get_code_from_dict(val)
                        if c:
                            return c
                elif isinstance(val, (str, int)):
                    c = extract_code_from_str(val)
                    if c:
                        return c

        if matches_target(res):
            c = get_code_from_dict(res)
            if c:
                return c

    return None

async def async_voltx_get_number(session, range_id, api_source="API1"):
    if api_source == "API2":
        api_key = get_setting('number_api_key_2') or "M6SB7HZXXIX"
        url = get_setting('getnum_url_2') or "https://api.2oo9.cloud/MXS47FLFX0U/tnevs/@public/api/getnum"
    else:
        api_key = get_setting('number_api_key') or "M7D4REK5Y06"
        url = get_setting('getnum_url') or "https://api.2oo9.cloud/MXS47FLFX0U/tnevs/@public/api/getnum"
        
    headers = {
        "mauthapi": api_key,
        "Content-Type": "application/json"
    }
    payload = {"rid": str(range_id)}
    
    try:
        async with session.post(url, json=payload, headers=headers, timeout=6) as response:
            if response.status == 200:
                res = await response.json()
                meta = res.get("meta", {})
                code = meta.get("code")
                
                if code == 200 and res.get("data"):
                    data = res["data"]
                    phone_num = data.get("full_number") or data.get("no_plus_number") or data.get("number")
                    order_id = data.get("no_plus_number") or phone_num
                    return True, order_id, phone_num, api_source
                elif code == 2946 or meta.get("status") == "not_found":
                    return False, "❌ নম্বর স্টকে নেই!", None, api_source
                else:
                    msg = res.get("message") or meta.get("status") or "API Error"
                    return False, f"⚠️ API Error: {msg}", None, api_source
            return False, "⚠️ API Server Error", None, api_source
    except Exception as e:
        return False, f"⚠️ সংযোগ বিচ্ছিন্ন: {e}", None, api_source

async def async_voltx_check_sms(session, phone_num, order_id, api_source="API1"):
    if api_source == "API2":
        api_key = get_setting('number_api_key_2') or "M6SB7HZXXIX"
        base_url = get_setting('api_base_url_2') or "https://api.2oo9.cloud/MXS47FLFX0U/tnevs/@public/api"
        getmsg_setting = get_setting('getmsg_url_2')
    else:
        api_key = get_setting('number_api_key') or "M7D4REK5Y06"
        base_url = get_setting('api_base_url') or "https://api.2oo9.cloud/MXS47FLFX0U/tnevs/@public/api"
        getmsg_setting = get_setting('getmsg_url')
    
    headers = {
        "mauthapi": api_key,
        "Content-Type": "application/json"
    }
    
    clean_phone = str(phone_num).replace("+", "").strip()
    clean_order_id = str(order_id).replace("+", "").strip() if order_id else clean_phone
    
    endpoints = []
    if getmsg_setting and getmsg_setting.strip():
        endpoints.append(getmsg_setting.strip())
    
    default_endpoints = [
        f"{base_url.rstrip('/')}/success-otp",
        f"{base_url.rstrip('/')}/getmsg"
    ]
    for ep in default_endpoints:
        if ep not in endpoints:
            endpoints.append(ep)
            
    for url in endpoints:
        if not url:
            continue
        clean_url = url.split('?')[0]
        
        try:
            payload = {
                "number": clean_phone,
                "no_plus_number": clean_phone,
                "full_number": f"+{clean_phone}",
                "id": clean_order_id
            }
            async with session.post(clean_url, json=payload, headers=headers, timeout=5) as r:
                if r.status == 200:
                    json_res = await r.json()
                    code = extract_otp_from_response(json_res, clean_phone, clean_order_id)
                    if code:
                        return "RECEIVED", code
        except Exception:
            pass

        for p_key, p_val in [('number', clean_phone), ('id', clean_order_id), ('phone', clean_phone)]:
            try:
                async with session.get(f"{clean_url}?{p_key}={p_val}", headers=headers, timeout=5) as r:
                    if r.status == 200:
                        json_res = await r.json()
                        code = extract_otp_from_response(json_res, clean_phone, clean_order_id)
                        if code:
                            return "RECEIVED", code
            except Exception:
                pass

    return "WAITING", None

# ----------------- 🔄 HIGH-SPEED ASYNC AUTO OTP POLLING THREAD -----------------
async def async_check_single_order(session, order):
    db_id, user_id, order_id, phone_num, service_name, c_code, api_source = order
    status, otp_code = await async_voltx_check_sms(session, phone_num, order_id, api_source=api_source)

    if status == "RECEIVED" and otp_code:
        otp_reward_val = float(get_setting('otp_reward') or 0.10)
        
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE active_orders SET status='COMPLETED', last_code=? WHERE id=?", (otp_code, db_id))
            cursor.execute("UPDATE users SET balance = balance + ?, reward_balance = reward_balance + ?, otp_count = otp_count + 1 WHERE user_id=?", (otp_reward_val, otp_reward_val, user_id))
            cursor.execute("SELECT balance, reward_balance FROM users WHERE user_id=?", (user_id,))
            row_user = cursor.fetchone()
            new_bal = row_user[0] if row_user else 0.0
            new_rew_bal = row_user[1] if row_user and len(row_user) > 1 else 0.0
            
            cursor.execute("SELECT country_flag FROM countries WHERE service_name=? AND country_code=? LIMIT 1", (service_name, c_code))
            flag_row = cursor.fetchone()
            c_flag = flag_row[0] if flag_row else "🇺🇿"
            conn.commit()

        backup_db_to_telegram()

        user_text = (
            f"🎉 <b>NEW OTP RECEIVED!</b>\n"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"{c_flag} <b>Number:</b> <code>{phone_num}</code>\n"
            f"📘 <b>OTP Code:</b> <code>{otp_code}</code>\n"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"🎁 <b>OTP Reward Credited: +{otp_reward_val:.2f} BDT</b>\n"
            f"💰 <b>Reward Balance: {new_rew_bal:.2f} BDT</b>\n"
            f"💼 <b>Total Balance: {new_bal:.2f} BDT</b>\n\n"
            f"👉 <i>(কোডের ওপর টাচ করলেই অটোমেটিক কপি হয়ে যাবে!)</i>"
        )
        try:
            bot.send_message(user_id, user_text)
        except Exception:
            pass

        otp_group = get_setting('otp_group_id') or "@otpreciverpro"
        group_text = (
            f"🔔 <b>LIVE OTP TRAFFIC!</b>\n"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"⚙️ <b>Service:</b> {service_name}\n"
            f"{c_flag} <b>Number:</b> <code>{phone_num}</code>\n"
            f"🔑 <b>OTP Code:</b> <code>{otp_code}</code>\n"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"🤖 <b>Bot:</b> @{BOT_USERNAME}"
        )
        try:
            bot.send_message(otp_group, group_text)
        except Exception as e:
            print(f"Group broadcast error: {e}")

    elif status == "CANCELLED":
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE active_orders SET status='CANCELLED' WHERE id=?", (db_id,))
            conn.commit()

async def async_otp_checker_worker():
    print("🚀 High-Speed Asynchronous OTP Checker Loop Started...")
    async with aiohttp.ClientSession() as session:
        while True:
            try:
                with get_db() as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT id, user_id, order_id, phone_number, service, country, api_source FROM active_orders WHERE status='WAITING'")
                    active_orders = cursor.fetchall()

                if active_orders:
                    tasks = [async_check_single_order(session, order) for order in active_orders]
                    await asyncio.gather(*tasks)

            except Exception as e:
                print(f"Error in Async OTP Worker: {e}")

            await asyncio.sleep(2)

def start_async_loop():
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(async_otp_checker_worker())

threading.Thread(target=start_async_loop, daemon=True).start()

# ----------------- MAIN KEYBOARD -----------------
def main_reply_keyboard(user_id):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    markup.add(types.KeyboardButton("📱 Get Free Number"), types.KeyboardButton("🛍️ Web Shop"))
    markup.add(types.KeyboardButton("👤 My Profile"), types.KeyboardButton("💳 Deposit"))
    markup.add(types.KeyboardButton("💸 Withdraw"), types.KeyboardButton("🔑 Get Code"))
    markup.add(types.KeyboardButton("🚥 LIVE TRAFFIC"), types.KeyboardButton("🎧 Support"))
    if user_id == ADMIN_ID:
        markup.add(types.KeyboardButton("👑 Admin Panel"))
    return markup

# ----------------- START COMMAND -----------------
@bot.message_handler(commands=['start'])
def start_cmd(message):
    user_id = message.from_user.id
    
    if not is_user_joined(user_id):
        send_force_join_msg(message.chat.id)
        return

    args = message.text.split()
    referred_by = int(args[1]) if len(args) > 1 and args[1].isdigit() and int(args[1]) != user_id else None

    signup_bonus = float(get_setting('signup_bonus') or 0.00)

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT user_id FROM users WHERE user_id=?", (user_id,))
        user = cursor.fetchone()

        if not user:
            cursor.execute("INSERT INTO users (user_id, balance, referred_by) VALUES (?, ?, ?)", (user_id, signup_bonus, referred_by))
            if referred_by:
                cursor.execute("UPDATE users SET referrals = referrals + 1 WHERE user_id=?", (referred_by,))
            conn.commit()
            backup_db_to_telegram()

    bot.send_message(message.chat.id, f"👋 <b>{BOT_NAME}</b>-এ আপনাকে স্বাগতম!", reply_markup=main_reply_keyboard(user_id))

# ----------------- 📱 AUTOMATED DUAL API BATCH NUMBER GETTER -----------------
@bot.message_handler(func=lambda msg: msg.text in ["📱 Get Number", "📱 Get Free Number", "📱 NUMBER'S"])
def numbers_cmd(message):
    if not is_user_joined(message.from_user.id):
        send_force_join_msg(message.chat.id)
        return

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT service_name FROM countries")
        services = cursor.fetchall()

    if not services:
        bot.send_message(message.chat.id, "⚠️ বর্তমানে কোনো সার্ভিস এভেলেবল নেই। এডমিন প্যানেল থেকে যোগ করুন।")
        return

    markup = types.InlineKeyboardMarkup(row_width=2)
    for s in services:
        srv_name = s[0]
        markup.add(types.InlineKeyboardButton(f"⚙️ {srv_name}", callback_data=f"usr_srv|{srv_name}"))

    bot.send_message(message.chat.id, "⚙️ <b>Select a Service for OTP:</b>", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("usr_srv|"))
def user_service_click(call):
    if not is_user_joined(call.from_user.id):
        bot.answer_callback_query(call.id, "⚠️ আগে আমাদের চ্যানেলে জয়েন করুন!", show_alert=True)
        send_force_join_msg(call.message.chat.id)
        return

    bot.answer_callback_query(call.id)
    parts = call.data.split("|")
    service_name = parts[1]
    
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT country_name, country_flag, country_code, service_code FROM countries WHERE service_name=?", (service_name,))
        countries = cursor.fetchall()
        
        markup = types.InlineKeyboardMarkup(row_width=1)
        for c in countries:
            c_name, c_flag, c_code, s_code = c
            btn_text = f"{c_flag} {c_name} (Dual API)"
            markup.add(types.InlineKeyboardButton(btn_text, callback_data=f"buy|{s_code}|{c_code}|{service_name}"))

    markup.add(types.InlineKeyboardButton("⬅️ Back To Services", callback_data="back_to_services"))
    bot.edit_message_text(f"🌍 <b>Select country for {service_name}:</b> ⬇️", call.message.chat.id, call.message.message_id, reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data == "back_to_services")
def back_to_services_cb(call):
    bot.answer_callback_query(call.id)
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT service_name FROM countries")
        services = cursor.fetchall()

    markup = types.InlineKeyboardMarkup(row_width=2)
    for s in services:
        srv_name = s[0]
        markup.add(types.InlineKeyboardButton(f"⚙️ {srv_name}", callback_data=f"usr_srv|{srv_name}"))

    bot.edit_message_text("⚙️ <b>Select a Service:</b>", call.message.chat.id, call.message.message_id, reply_markup=markup)

async def async_fetch_batch_numbers(range_id):
    async with aiohttp.ClientSession() as session:
        tasks = [
            async_voltx_get_number(session, range_id, "API1"),
            async_voltx_get_number(session, range_id, "API1"),
            async_voltx_get_number(session, range_id, "API2"),
            async_voltx_get_number(session, range_id, "API2")
        ]
        results = await asyncio.gather(*tasks)
        return results

@bot.callback_query_handler(func=lambda call: call.data.startswith("buy|"))
def buy_number_click(call):
    if not is_user_joined(call.from_user.id):
        bot.answer_callback_query(call.id, "⚠️ আগে আমাদের চ্যানেলে জয়েন করুন!", show_alert=True)
        send_force_join_msg(call.message.chat.id)
        return

    bot.answer_callback_query(call.id, "⌛ DUAL API (VoltXSMS + StexSMS) থেকে ৪টি নম্বর আনা হচ্ছে...", show_alert=False)
    
    parts = call.data.split("|")
    range_id = parts[1]
    c_code = parts[2]
    service_name = parts[3]
    user_id = call.from_user.id

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT country_name, country_flag, price FROM countries WHERE service_name=? AND country_code=? LIMIT 1", (service_name, c_code))
        c_row = cursor.fetchone()
        
        c_name = c_row[0] if c_row else "Uzbekistan"
        c_flag = c_row[1] if c_row else "🇺🇿"

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    results = loop.run_until_complete(async_fetch_batch_numbers(range_id))
    loop.close()

    assigned_numbers = []
    with get_db() as conn:
        cursor = conn.cursor()
        for success, order_id_or_err, phone_num, api_src in results:
            if success and phone_num:
                assigned_numbers.append(phone_num)
                cursor.execute(
                    "INSERT INTO active_orders (user_id, order_id, phone_number, service, country, api_source) VALUES (?, ?, ?, ?, ?, ?)",
                    (user_id, order_id_or_err, phone_num, service_name, c_code, api_src)
                )
        conn.commit()

    otp_group_link = get_setting('otp_group_link') or 'https://t.me/otpreciverpro'
    otp_reward_val = float(get_setting('otp_reward') or 0.10)

    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(types.InlineKeyboardButton("👀 OTP GROUP ↗️", url=otp_group_link))
    markup.add(
        types.InlineKeyboardButton("⚙️ Next Number", callback_data=f"buy|{range_id}|{c_code}|{service_name}"),
        types.InlineKeyboardButton("🌐 Select Country", callback_data=f"usr_srv|{service_name}")
    )

    if not assigned_numbers:
        err_text = (
            f"❌ <b>নম্বর আনা সম্ভব হয়নি!</b>\n\n"
            f"বর্তমানে {c_flag} <b>{c_name} ({service_name})</b> এর জন্য DUAL API-তে নম্বর স্টকে নেই।\n\n"
            f"👇 <i>আবার চেষ্টা করতে বা কান্ট্রি চেঞ্জ করতে নিচের বাটন ব্যবহার করুন:</i>"
        )
        try:
            bot.edit_message_text(err_text, call.message.chat.id, call.message.message_id, reply_markup=markup)
        except Exception:
            bot.send_message(call.message.chat.id, err_text, reply_markup=markup)
        return

    num_str_list = "\n".join([f"{c_flag} 📋 <code>{p}</code>" for p in assigned_numbers])

    text = (
        f"{c_flag} <b>{c_name}</b> 📘 <b>Number Assigned (Dual API)</b>\n\n"
        f"💰 <b>Per OTP Reward : {otp_reward_val:.2f} BDT</b>\n"
        f"━━━━━━━━━━━━━━━━━━━\n"
        f"{num_str_list}\n\n"
        f"⏳ <i>Waiting for OTP... (অটোমেটিক চেক হচ্ছে)</i>"
    )

    try:
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=markup)
    except Exception:
        bot.send_message(call.message.chat.id, text, reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data == "dummy_copy")
def dummy_copy_cb(call):
    bot.answer_callback_query(call.id, "📋 নম্বর বা কোডের ওপর ক্লিক করুন, অটো কপি হয়ে যাবে!", show_alert=False)

@bot.message_handler(func=lambda msg: msg.text == "🚥 LIVE TRAFFIC")
def live_traffic_cmd(message):
    bot.send_message(message.chat.id, "📊 <b>LIVE TRAFFIC</b>\n━━━━━━━━━━━━━━━━━━\n🌐 <b>SMS Engine:</b> Async DUAL API Connected (VoltX + Stex)\nOTP Monitoring Running 24/7 Auto Polling...")

# ----------------- 🛍️ WEB SHOP SYSTEM -----------------
@bot.message_handler(func=lambda msg: msg.text == "🛍️ Web Shop")
def buy_products_cmd(message):
    if not is_user_joined(message.from_user.id):
        send_force_join_msg(message.chat.id)
        return

    text = "🛍️ <b>Buy Products</b>\n\nSelect a category:"
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        types.InlineKeyboardButton("🛡️ VPN", callback_data="cat_VPN"),
        types.InlineKeyboardButton("🌐 Proxy", callback_data="cat_Proxy")
    )
    markup.add(types.InlineKeyboardButton("📧 Mail", callback_data="cat_Mail"))
    bot.send_message(message.chat.id, text, reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("cat_"))
def category_select_cb(call):
    bot.answer_callback_query(call.id)
    category = call.data.split("_")[1]
    
    if category == "VPN":
        text = "🛡️ <b>VPN — Select Duration:</b>"
        markup = types.InlineKeyboardMarkup(row_width=2)
        markup.add(
            types.InlineKeyboardButton("🔐 3 Days", callback_data="vpndur_3 Days"),
            types.InlineKeyboardButton("🔐 7 Days", callback_data="vpndur_7 Days")
        )
        markup.add(
            types.InlineKeyboardButton("🔐 14 Days", callback_data="vpndur_14 Days"),
            types.InlineKeyboardButton("🔐 30 Days", callback_data="vpndur_30 Days")
        )
        markup.add(types.InlineKeyboardButton("‹ Back", callback_data="back_to_shop"))
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=markup)
        return

    elif category == "Proxy":
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id, name, price, subcategory FROM products WHERE category='Proxy'")
            products = cursor.fetchall()
            
            if not products:
                bot.answer_callback_query(call.id, "⚠️ প্রক্সি ক্যাটাগরিতে বর্তমানে কোনো প্ল্যান নেই!", show_alert=True)
                return

            text = "🌐 <b>Proxy — Select Plan:</b>"
            markup = types.InlineKeyboardMarkup(row_width=1)
            
            for p in products:
                p_id, p_name, p_price, p_sub = p
                cursor.execute("SELECT COUNT(*) FROM item_stock WHERE product_id=? AND status='available'", (p_id,))
                p_stock = cursor.fetchone()[0]
                plan_info = f" ({p_sub})" if p_sub else ""
                btn_text = f"🌐 {p_name} · {p_price:.2f} BDT/unit{plan_info}"
                markup.add(types.InlineKeyboardButton(btn_text, callback_data=f"selectprod_{p_id}"))
                
            markup.add(types.InlineKeyboardButton("‹ Back", callback_data="back_to_shop"))
            bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=markup)
            return

    else:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id, name, price FROM products WHERE category='Mail'")
            products = cursor.fetchall()
            
            if not products:
                bot.answer_callback_query(call.id, "⚠️ মেইল ক্যাটাগরিতে বর্তমানে কোনো প্রোডাক্ট নেই!", show_alert=True)
                return

            text = f"<b>Mail — Select Product:</b>"
            markup = types.InlineKeyboardMarkup(row_width=1)
            
            for p in products:
                p_id, p_name, p_price = p
                cursor.execute("SELECT COUNT(*) FROM item_stock WHERE product_id=? AND status='available'", (p_id,))
                p_stock = cursor.fetchone()[0]
                btn_text = f"📧 {p_name} · {p_price:.2f} BDT · {p_stock} in stock"
                markup.add(types.InlineKeyboardButton(btn_text, callback_data=f"selectprod_{p_id}"))
                
        markup.add(types.InlineKeyboardButton("‹ Back", callback_data="back_to_shop"))
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("vpndur_"))
def vpn_duration_select_cb(call):
    bot.answer_callback_query(call.id)
    duration = call.data.split("_")[1]
    
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id, name, price FROM products WHERE category='VPN' AND subcategory=?", (duration,))
        products = cursor.fetchall()
        
        if not products:
            bot.answer_callback_query(call.id, f"⚠️ {duration} এর জন্য বর্তমানে কোনো VPN নেই!", show_alert=True)
            return

        text = f"🛡️ <b>VPN — {duration} Packages:</b>"
        markup = types.InlineKeyboardMarkup(row_width=1)
        for p in products:
            p_id, p_name, p_price = p
            cursor.execute("SELECT COUNT(*) FROM item_stock WHERE product_id=? AND status='available'", (p_id,))
            p_stock = cursor.fetchone()[0]
            btn_text = f"🔐 {p_name} ({duration}) · {p_price:.2f} BDT · {p_stock} in stock"
            markup.add(types.InlineKeyboardButton(btn_text, callback_data=f"selectprod_{p_id}"))
            
    markup.add(types.InlineKeyboardButton("‹ Back", callback_data="cat_VPN"))
    bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data == "back_to_shop")
def back_to_shop_cb(call):
    bot.answer_callback_query(call.id)
    text = "🛍️ <b>Buy Products</b>\n\nSelect a category:"
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        types.InlineKeyboardButton("🛡️ VPN", callback_data="cat_VPN"),
        types.InlineKeyboardButton("🌐 Proxy", callback_data="cat_Proxy")
    )
    markup.add(types.InlineKeyboardButton("📧 Mail", callback_data="cat_Mail"))
    bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("selectprod_"))
def select_prod_cb(call):
    bot.answer_callback_query(call.id)
    p_id = call.data.split("_")[1]
    user_id = call.from_user.id
    
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT name, price, category, subcategory FROM products WHERE id=?", (p_id,))
        prod = cursor.fetchone()
        
        if not prod:
            return
            
        p_name, p_price, p_cat, p_sub = prod
        cursor.execute("SELECT COUNT(*) FROM item_stock WHERE product_id=? AND status='available'", (p_id,))
        p_stock = cursor.fetchone()[0]
    
    set_user_state(user_id, "buying_p_id", str(p_id))
    set_user_state(user_id, "step", "await_qty")
    
    sub_info = f" ({p_sub})" if p_sub else ""
    text = (
        f"📦 <b>{p_name}{sub_info}</b>\n"
        f"⚡ <b>Price: {p_price:.2f} BDT / piece</b>\n"
        f"⚡ <b>Available: {p_stock}</b>\n\n"
        f"Enter quantity:"
    )
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_action"))
    bot.send_message(call.message.chat.id, text, reply_markup=markup)

# ----------------- ORDER CONFIRMATION -----------------
@bot.callback_query_handler(func=lambda call: call.data.startswith("cfmbuy_"))
def confirm_order_cb(call):
    bot.answer_callback_query(call.id)
    user_id = call.from_user.id
    
    try:
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=None)
    except Exception:
        pass

    parts = call.data.split("_")
    p_id = int(parts[1])
    qty = int(parts[2])
    
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT name, price, category, subcategory FROM products WHERE id=?", (p_id,))
        prod = cursor.fetchone()
        if not prod:
            bot.edit_message_text("❌ প্রোডাক্টটি পাওয়া যায়নি!", call.message.chat.id, call.message.message_id)
            return
            
        p_name, p_price, p_cat, p_sub = prod
        total_bdt = qty * p_price
        
        cursor.execute("SELECT balance FROM users WHERE user_id=?", (user_id,))
        user_row = cursor.fetchone()
        bal = user_row[0] if user_row else 0.0
        
        if bal < total_bdt:
            bot.edit_message_text(
                f"❌ <b>অর্ডার সম্পূর্ণ করা সম্ভব হয়নি!</b>\n\n"
                f"আপনার ব্যালেন্স: <b>{bal:.2f} BDT</b>\n"
                f"প্রয়োজন: <b>{total_bdt:.2f} BDT</b>\n\n"
                f"অনুগ্রহ করে 💳 Deposit সার্ভিস থেকে রিচার্জ করুন।",
                call.message.chat.id, call.message.message_id
            )
            return
            
        cursor.execute("SELECT id, content FROM item_stock WHERE product_id=? AND status='available' LIMIT ?", (p_id, qty))
        items = cursor.fetchall()
        
        if len(items) < qty:
            bot.edit_message_text("❌ পর্যাপ্ত স্টক খালি নেই!", call.message.chat.id, call.message.message_id)
            return
            
        delivered_lines = []
        for item in items:
            item_id, content = item
            delivered_lines.append(content)
            cursor.execute("UPDATE item_stock SET status='sold' WHERE id=?", (item_id,))
            
        cursor.execute("UPDATE users SET balance = balance - ? WHERE user_id=?", (total_bdt, user_id))
        
        joined_content = "\n".join(delivered_lines)
        cursor.execute(
            "INSERT INTO purchases (user_id, product_name, category, qty, total_price, content) VALUES (?, ?, ?, ?, ?, ?)",
            (user_id, p_name, p_cat, qty, total_bdt, joined_content)
        )
        conn.commit()

    backup_db_to_telegram()

    if p_cat == "Proxy":
        formatted_proxies = []
        for idx, line in enumerate(delivered_lines, 1):
            parts = line.split(":")
            if len(parts) == 4:
                ip, port, usr, pwd = parts
                formatted_proxies.append(
                    f"🌐 <b>Proxy #{idx}</b>\n"
                    f"🖥️ <code>{ip}</code>\n"
                    f"🔌 <code>{port}</code>\n"
                    f"👤 <code>{usr}</code>\n"
                    f"🔓 <code>{pwd}</code>"
                )
            else:
                formatted_proxies.append(f"🌐 <b>Proxy #{idx}:</b> <code>{line}</code>")

        msg1_text = (
            f"✅ <b>Proxy Delivered!</b>\n\n"
            f"⚡ <b>{p_name}</b>\n"
            f"_______________________\n\n" +
            "\n\n".join(formatted_proxies)
        )
        bot.edit_message_text(msg1_text, call.message.chat.id, call.message.message_id)

        sub_info = f" ({p_sub})" if p_sub else ""
        msg2_text = (
            f"⚡ <b>{p_name} — × {qty} units{sub_info}</b>\n"
            f"💰 <b>{total_bdt:.2f} BDT</b>\n"
            f"_______________________\n"
            f"📩 <b>Credentials sent above ↑</b>"
        )
        bot.send_message(call.message.chat.id, msg2_text)

    elif p_cat == "VPN":
        formatted_vpns = []
        for idx, line in enumerate(delivered_lines, 1):
            if ":" in line:
                email, pwd = line.split(":", 1)
                formatted_vpns.append(
                    f"🛡️ <b>VPN #{idx}</b>\n"
                    f"🖥️ <code>{email}</code>\n"
                    f"🔓 <code>{pwd}</code>"
                )
            elif " | " in line:
                email, pwd = line.split(" | ", 1)
                formatted_vpns.append(
                    f"🛡️ <b>VPN #{idx}</b>\n"
                    f"🖥️ <code>{email}</code>\n"
                    f"🔓 <code>{pwd}</code>"
                )
            else:
                formatted_vpns.append(f"🛡️ <b>VPN #{idx}:</b> <code>{line}</code>")

        msg1_text = (
            f"✅ <b>VPN Delivered!</b>\n\n"
            f"⚡ <b>{p_name}</b>\n"
            f"_______________________\n\n" +
            "\n\n".join(formatted_vpns)
        )
        bot.edit_message_text(msg1_text, call.message.chat.id, call.message.message_id)

        sub_info = f" ({p_sub})" if p_sub else ""
        msg2_text = (
            f"⚡ <b>{p_name} — × {qty} units{sub_info}</b>\n"
            f"💰 <b>{total_bdt:.2f} BDT</b>\n"
            f"_______________________\n"
            f"📩 <b>Credentials sent above ↑</b>"
        )
        bot.send_message(call.message.chat.id, msg2_text)

    else:
        excel_file = create_excel_document(p_name, delivered_lines)
        
        delivery_text = (
            f"✅ <b>Mail Delivered!</b>\n\n"
            f"📧 <b>{qty}x {p_name}</b>\n"
            f"💰 <b>Paid : {total_bdt:.2f} BDT</b>\n"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"📱 <b>File below ↓</b>"
        )
        bot.edit_message_text(delivery_text, call.message.chat.id, call.message.message_id)
        bot.send_document(call.message.chat.id, excel_file)

# ----------------- 👤 PROFILE & PURCHASE HISTORY -----------------
@bot.message_handler(func=lambda msg: msg.text == "👤 My Profile")
def profile_cmd(message):
    if not is_user_joined(message.from_user.id):
        send_force_join_msg(message.chat.id)
        return

    user_id = message.from_user.id
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT balance, reward_balance, referrals, joined_date, otp_count FROM users WHERE user_id=?", (user_id,))
        row = cursor.fetchone()

    bal = row[0] if row else 0.00
    rew_bal = row[1] if row and len(row) > 1 else 0.00
    refs = row[2] if row and len(row) > 2 else 0
    joined = row[3] if row and len(row) > 3 else "2026-06-27"
    otps = row[4] if row and len(row) > 4 else 0
    usdt = bal / 125.0
    ref_link = f"https://t.me/{BOT_USERNAME}?start={user_id}"

    text = (
        f"👤 <b>My Profile</b>\n\n"
        f"👤 <b>Name     :</b> {message.from_user.first_name}\n"
        f"🆔 <b>User ID  :</b> <code>{user_id}</code>\n"
        f"💰 <b>Balance  :</b> {bal:.2f} BDT / {usdt:.4f} USDT\n"
        f"🎁 <b>Reward Bal :</b> {rew_bal:.2f} BDT <i>(Withdrawable)</i>\n"
        f"🥳 <b>Joined   :</b> {joined}\n"
        f"🔢 <b>OTP Received :</b> {otps}\n"
        f"🤝 <b>Referrals :</b> {refs}\n\n"
        f"🔗 <b>Referral Link:</b>\n<code>{ref_link}</code>"
    )
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("📜 Purchase History", callback_data="view_purchases"))
    bot.send_message(message.chat.id, text, reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data == "view_purchases")
def view_purchases_cb(call):
    bot.answer_callback_query(call.id)
    user_id = call.from_user.id
    
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT product_name, category, total_price, content, date FROM purchases WHERE user_id=? ORDER BY id DESC LIMIT 5", (user_id,))
        purchases = cursor.fetchall()

    if not purchases:
        bot.send_message(call.message.chat.id, "📜 <b>আপনার কোনো পূর্ববর্তী পারচেজ হিস্ট্রি নেই।</b>")
        return

    for p in purchases:
        p_name, p_cat, p_price, p_content, p_date = p
        p_type = f"{p_cat.upper()}_QTY"
        
        hist_text = (
            f"⚡ <b>Purchase Details</b>\n\n"
            f"<b>Type</b>     : {p_type}\n"
            f"<b>Date</b>     : {p_date}\n"
            f"<b>Total</b>    : {p_price:.2f} BDT\n\n"
            f"<b>Credentials:</b>\n<code>{p_content}</code>"
        )
        bot.send_message(call.message.chat.id, hist_text)

# ----------------- 💸 WITHDRAWAL SYSTEM (BKASH, NAGAD, BINANCE) -----------------
@bot.message_handler(func=lambda msg: msg.text in ["💸 Withdraw", "Withdraw"])
def withdraw_cmd(message):
    user_id = message.from_user.id
    if not is_user_joined(user_id):
        send_force_join_msg(message.chat.id)
        return

    min_w = float(get_setting('min_withdraw') or 50.0)
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT balance, reward_balance FROM users WHERE user_id=?", (user_id,))
        row = cursor.fetchone()
        bal = row[0] if row else 0.0
        rew_bal = row[1] if row and len(row) > 1 else 0.0

    if rew_bal < min_w:
        bot.send_message(
            message.chat.id,
            f"❌ <b>উইথড্র করার জন্য পর্যাপ্ত রিওয়ার্ড ব্যালেন্স নেই!</b>\n\n"
            f"আপনার রিওয়ার্ড ব্যালেন্স: <b>{rew_bal:.2f} BDT</b>\n"
            f"(ডিপোজিট করা ব্যালেন্স দিয়ে উইথড্র করা যাবে না)\n"
            f"সর্বনিম্ন উইথড্র: <b>{min_w:.2f} BDT</b>"
        )
        return

    text = "💸 <b>Select Withdrawal Method:</b>"
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        types.InlineKeyboardButton("🌸 bKash", callback_data="w_method_bkash"),
        types.InlineKeyboardButton("🟠 Nagad", callback_data="w_method_nagad")
    )
    markup.add(types.InlineKeyboardButton("🟡 Binance (USDT)", callback_data="w_method_binance"))
    bot.send_message(message.chat.id, text, reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("w_method_"))
def withdraw_method_cb(call):
    bot.answer_callback_query(call.id)
    method = call.data.split("_")[2].upper()
    user_id = call.from_user.id

    set_user_state(user_id, "w_method", method)
    set_user_state(user_id, "step", "await_w_acc")

    text = f"✏️ <b>Enter your {method} Account Number / Pay UID:</b>"
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_action"))
    bot.send_message(call.message.chat.id, text, reply_markup=markup)

# ----------------- 💳 DEPOSIT SYSTEM -----------------
@bot.message_handler(func=lambda msg: msg.text == "💳 Deposit")
def deposit_cmd(message):
    if not is_user_joined(message.from_user.id):
        send_force_join_msg(message.chat.id)
        return

    text = (
        "💳 <b>Select Deposit Payment Method:</b>\n\n"
        "নিচের তালিকা থেকে আপনার সুবিধাজনক পেমেন্ট গেটওয়ে সিলেক্ট করুন:"
    )
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        types.InlineKeyboardButton("💸 bKash (Personal)", callback_data="dep_bkash"),
        types.InlineKeyboardButton("💸 Nagad (Personal)", callback_data="dep_nagad")
    )
    markup.add(
        types.InlineKeyboardButton("💸 Binance (Pay / UID)", callback_data="dep_binance"),
        types.InlineKeyboardButton("🎟️ Cash Voucher", callback_data="dep_voucher")
    )
    bot.send_message(message.chat.id, text, reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("dep_"))
def dep_method_cb(call):
    bot.answer_callback_query(call.id)
    method = call.data.split("_")[1]
    user_id = call.from_user.id
    
    if method in ["bkash", "nagad", "binance"]:
        min_dep = get_setting('min_deposit') or '20.0'
        num = get_setting(f'{method}_num') if method != 'binance' else get_setting('binance_uid')
        
        set_user_state(user_id, "dep_method", method)
        set_user_state(user_id, "dep_num", num)
        set_user_state(user_id, "step", "await_dep_amt")
        
        header_badge = "💸 <b>[ bKash Personal Payment ]</b>" if method == "bkash" else ("💸 <b>[ Nagad Personal Payment ]</b>" if method == "nagad" else "💸 <b>[ Binance USDT Payment ]</b>")
        
        text = f"{header_badge}\n\nEnter deposit amount in BDT:\n<i>(Minimum: {min_dep} BDT)</i>"
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_action"))
        bot.send_message(call.message.chat.id, text, reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("paid_"))
def dep_paid_cb(call):
    bot.answer_callback_query(call.id)
    user_id = call.from_user.id
    _, method, amt = call.data.split("_")
    
    set_user_state(user_id, "dep_final_m", method)
    set_user_state(user_id, "dep_final_a", amt)
    set_user_state(user_id, "step", "await_trxid")
    
    text = (
        f"🚩 <b>Enter Transaction ID (TrxID)</b>\n\n"
        f"Amount: <b>{float(amt):.2f} BDT</b> via <b>💸 {method.upper()}</b>\n\n"
        f"Send the TrxID from your payment SMS below:\n"
        f"<i>(উদাহরণ: DF27TNVV17)</i>"
    )
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_action"))
    bot.send_message(call.message.chat.id, text, reply_markup=markup)

# DIRECT DONGVANFB MAILBOX READER LINK
@bot.message_handler(func=lambda msg: msg.text == "🔑 Get Code")
def get_code_cmd(message):
    text = "🔑 <b>Get Code</b>\n\nSelect a link to open mailbox reader:"
    markup = types.InlineKeyboardMarkup(row_width=1)
    markup.add(
        types.InlineKeyboardButton("🔗 Hotmail/Outlook ↗", url="https://dongvanfb.net/read_mail_box/"),
        types.InlineKeyboardButton("🔗 Fr Outlook Code ↗", url="https://dongvanfb.net/read_mail_box/"),
        types.InlineKeyboardButton("🔗 API Gmail Code ↗", url="https://dongvanfb.net/read_mail_box/")
    )
    bot.send_message(message.chat.id, text, reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data == "cancel_action")
def cancel_action_cb(call):
    bot.answer_callback_query(call.id)
    user_id = call.from_user.id
    clear_user_state(user_id)
    bot.edit_message_text("❌ অপশনটি বাতিল করা হয়েছে।", call.message.chat.id, call.message.message_id)

@bot.message_handler(func=lambda msg: msg.text == "🎧 Support")
def support_cmd(message):
    bot.send_message(message.chat.id, "🎧 <b>Support:</b>\n\nযেকোনো সমস্যায় এডমিনের সাথে যোগাযোগ করুন: @your_telegram_username")

# ----------------- 👑 EDITABLE ADMIN CONTROL PANEL -----------------
@bot.message_handler(func=lambda msg: msg.text == "👑 Admin Panel" and msg.from_user.id == ADMIN_ID)
def admin_panel_cmd(message):
    text = "📊 <b>MASTER ADMIN CONTROL PANEL</b>\n\nসবকিছু কাস্টমাইজ করতে নিচের অপশনগুলো ব্যবহার করুন:"
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        types.InlineKeyboardButton("⚙️ VoltX (API 1) Config", callback_data="adm_voltx_config"),
        types.InlineKeyboardButton("⚙️ StexSMS (API 2) Config", callback_data="adm_stex_config")
    )
    markup.add(
        types.InlineKeyboardButton("📢 Force Join Channels", callback_data="adm_edit_force_join"),
        types.InlineKeyboardButton("💸 Edit Payment Numbers", callback_data="adm_edit_payments")
    )
    markup.add(
        types.InlineKeyboardButton("👀 Pending Deposits", callback_data="adm_view_pending_dep"),
        types.InlineKeyboardButton("📤 Pending Withdrawals", callback_data="adm_view_pending_w")
    )
    markup.add(
        types.InlineKeyboardButton("🎁 Edit Bonuses & Rewards", callback_data="adm_edit_bonuses"),
        types.InlineKeyboardButton("🌐 Add Service/Country", callback_data="adm_add_country")
    )
    markup.add(
        types.InlineKeyboardButton("🛍️ Add Shop Stock", callback_data="adm_add_shop_stock"),
        types.InlineKeyboardButton("👤 Edit User Balance", callback_data="adm_edit_balance")
    )
    markup.add(
        types.InlineKeyboardButton("📢 Broadcast Message", callback_data="adm_broadcast"),
        types.InlineKeyboardButton("🗑️ Reset & Clear All Stock", callback_data="adm_clear_stock")
    )
    bot.send_message(message.chat.id, text, reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.from_user.id == ADMIN_ID)
def admin_cb_handler(call):
    data = call.data
    bot.answer_callback_query(call.id)
    user_id = call.from_user.id

    if data == "adm_voltx_config":
        base_url = get_setting('api_base_url')
        api_key = get_setting('number_api_key')
        config_text = (
            f"👑 <b>API 1 (VoltXSMS Engine) Config</b>\n\n"
            f"🌐 <b>Base URL:</b> <code>{base_url}</code>\n"
            f"🔐 <b>API Key :</b> <code>{api_key}</code>\n"
        )
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            types.InlineKeyboardButton("Edit API 1 Base URL", callback_data="set_cfg_api_base_url"),
            types.InlineKeyboardButton("Edit API 1 Key", callback_data="set_cfg_number_api_key"),
            types.InlineKeyboardButton("‹ Back", callback_data="adm_back_main")
        )
        bot.edit_message_text(config_text, call.message.chat.id, call.message.message_id, reply_markup=markup)

    elif data == "adm_stex_config":
        base_url_2 = get_setting('api_base_url_2')
        api_key_2 = get_setting('number_api_key_2')
        config_text = (
            f"👑 <b>API 2 (StexSMS Engine) Config</b>\n\n"
            f"🌐 <b>Base URL:</b> <code>{base_url_2}</code>\n"
            f"🔐 <b>API Key :</b> <code>{api_key_2}</code>\n"
        )
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            types.InlineKeyboardButton("Edit API 2 Base URL", callback_data="set_cfg_api_base_url_2"),
            types.InlineKeyboardButton("Edit API 2 Key", callback_data="set_cfg_number_api_key_2"),
            types.InlineKeyboardButton("‹ Back", callback_data="adm_back_main")
        )
        bot.edit_message_text(config_text, call.message.chat.id, call.message.message_id, reply_markup=markup)

    elif data.startswith("set_cfg_"):
        key = data.replace("set_cfg_", "")
        set_user_state(user_id, "adm_step", f"set_setting:{key}")
        bot.send_message(call.message.chat.id, f"✏️ <b>নতুন {key} এর মান লিখে পাঠান:</b>")

    elif data == "adm_back_main":
        admin_panel_cmd(call.message)

    elif data == "adm_edit_bonuses":
        ref = get_setting('refer_reward')
        dep_b = get_setting('deposit_bonus')
        otp_b = get_setting('otp_reward')
        sgn_b = get_setting('signup_bonus') or "0.00"
        text = (
            f"🎁 <b>Edit Bonuses & Rewards</b>\n\n"
            f"Signup Bonus: {sgn_b} BDT\n"
            f"Refer Reward: {ref} BDT\n"
            f"Deposit Bonus: {dep_b}%\n"
            f"OTP Reward: {otp_b} BDT"
        )
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            types.InlineKeyboardButton("🎁 Change Signup Joining Bonus", callback_data="set_bonus_signup"),
            types.InlineKeyboardButton("🤝 Change Refer Reward", callback_data="set_bonus_refer"),
            types.InlineKeyboardButton("💵 Change Deposit Bonus %", callback_data="set_bonus_dep"),
            types.InlineKeyboardButton("🎉 Change Per-OTP Reward", callback_data="set_bonus_otpreward")
        )
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=markup)

    elif data == "set_bonus_signup":
        set_user_state(user_id, "adm_step", "set_setting:signup_bonus")
        bot.send_message(call.message.chat.id, "✏️ <b>নতুন ইউজার জয়েন করলে কত টাকা বোনাস দিতে চান লিখে পাঠান (যেমন: 0.00 বা 6.00):</b>")

    elif data == "set_bonus_otpreward":
        set_user_state(user_id, "adm_step", "set_setting:otp_reward")
        bot.send_message(call.message.chat.id, "✏️ <b>প্রতি OTP তে ইউজারকে কত টাকা রিওয়ার্ড দিতে চান লিখে পাঠান (যেমন: 0.10):</b>")

    elif data == "adm_view_pending_w":
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id, user_id, amount, method, account_number FROM withdrawals WHERE status='pending'")
            w_reqs = cursor.fetchall()
        
        if not w_reqs:
            bot.send_message(call.message.chat.id, "✅ বর্তমানে কোনো পেন্ডিং উইথড্র রিকুয়েস্ট নেই।")
            return
            
        for w in w_reqs:
            w_id, u_id, amt, method, acc = w
            t = f"📤 <b>Pending Withdrawal #{w_id}</b>\nUser: <code>{u_id}</code>\nMethod: <b>{method}</b>\nAccount: <code>{acc}</code>\nAmount: <b>{amt:.2f} BDT</b>"
            m = types.InlineKeyboardMarkup(row_width=2)
            m.add(
                types.InlineKeyboardButton("✅ Approve", callback_data=f"adm_appr_w_{w_id}"),
                types.InlineKeyboardButton("❌ Reject & Refund", callback_data=f"adm_rej_w_{w_id}")
            )
            bot.send_message(call.message.chat.id, t, reply_markup=m)

    elif data.startswith("adm_appr_w_"):
        w_id = data.split("_")[3]
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE withdrawals SET status='approved' WHERE id=?", (w_id,))
            conn.commit()
        backup_db_to_telegram()
        bot.edit_message_text(f"✅ <b>Withdrawal #{w_id} Approved & Paid!</b>", call.message.chat.id, call.message.message_id)

    elif data.startswith("adm_rej_w_"):
        w_id = data.split("_")[3]
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT user_id, amount FROM withdrawals WHERE id=? AND status='pending'", (w_id,))
            row = cursor.fetchone()
            if row:
                u_id, amt = row
                cursor.execute("UPDATE withdrawals SET status='rejected' WHERE id=?", (w_id,))
                cursor.execute("UPDATE users SET reward_balance = reward_balance + ? WHERE user_id=?", (amt, u_id))
                conn.commit()
                backup_db_to_telegram()
                bot.edit_message_text(f"❌ <b>Withdrawal #{w_id} Rejected & {amt:.2f} BDT Refunded to User Reward Balance!</b>", call.message.chat.id, call.message.message_id)

    elif data == "adm_edit_force_join":
        curr = get_setting('force_channels') or "None"
        set_user_state(user_id, "adm_step", "set_setting:force_channels")
        bot.send_message(
            call.message.chat.id,
            f"📢 <b>Current Required Channels:</b> <code>{curr}</code>\n\n"
            f"ইউজারদের বাধ্যতামূলক জয়েন করানোর জন্য চ্যানেলগুলোর ইউজারনেম কমা দিয়ে দিয়ে লিখুন:\n"
            f"উদাহরণ: <code>@channel1,@channel2</code>"
        )

    elif data == "adm_clear_stock":
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM countries")
            cursor.execute("DELETE FROM products")
            cursor.execute("DELETE FROM item_stock")
            conn.commit()
        backup_db_to_telegram()
        bot.send_message(call.message.chat.id, "🗑️ আগের সকল সার্ভিস ও ইনভেন্টরি ক্লিয়ার করা হয়েছে!")

    elif data == "adm_add_shop_stock":
        text = "🛍️ <b>Select Category to Add Product / Stock:</b>"
        markup = types.InlineKeyboardMarkup(row_width=3)
        markup.add(
            types.InlineKeyboardButton("📧 Mail Stock", callback_data="addstk_Mail"),
            types.InlineKeyboardButton("🌐 Proxy Stock", callback_data="addstk_Proxy"),
            types.InlineKeyboardButton("🛡️ VPN Stock", callback_data="addstk_VPN")
        )
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=markup)

    elif data.startswith("addstk_"):
        cat = data.split("_")[1]
        set_user_state(user_id, "adm_add_cat", cat)
        set_user_state(user_id, "adm_step", "add_prod_info")
        
        if cat == "VPN":
            instruct = (
                "🛡️ <b>VPN Product Setup:</b>\n\n"
                "নিচের ফরম্যাটে লিখুন:\n"
                "<code>মেয়াদ (3 Days / 7 Days / 14 Days / 30 Days), নাম, দাম</code>\n\n"
                "উদাহরণ:\n<code>3 Days, ExpressVPN, 25.00</code>"
            )
        elif cat == "Proxy":
            instruct = (
                "🌐 <b>Proxy Product Setup:</b>\n\n"
                "নিচের ফরম্যাটে লিখুন:\n"
                "<code>প্ল্যান/MB, নাম, দাম</code>\n\n"
                "উদাহরণ:\n<code>200 MB, Owl Proxy, 10.00</code>"
            )
        else:
            instruct = (
                "📧 <b>Mail Product Setup:</b>\n\n"
                "নিচের ফরম্যাটে লিখুন:\n"
                "<code>নাম, দাম</code>\n\n"
                "উদাহরণ:\n<code>Fr Outlook, 0.80</code>"
            )
        bot.send_message(call.message.chat.id, instruct)

    elif data == "adm_view_pending_dep":
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id, user_id, amount, method, trx_id FROM deposits WHERE status='pending'")
            deps = cursor.fetchall()
        
        if not deps:
            bot.send_message(call.message.chat.id, "✅ বর্তমানে কোনো পেন্ডিং ডিপোজিট রিকুয়েস্ট নেই।")
            return
            
        for d in deps:
            d_id, u_id, amt, method, trx = d
            t = f"📥 <b>Pending Deposit #{d_id}</b>\nUser: <code>{u_id}</code>\nMethod: {method}\nAmount: <b>{amt:.2f} BDT</b>\nTrxID: <code>{trx}</code>"
            m = types.InlineKeyboardMarkup(row_width=2)
            m.add(
                types.InlineKeyboardButton("✅ Approve", callback_data=f"adm_appr_dep_{d_id}"),
                types.InlineKeyboardButton("❌ Reject", callback_data=f"adm_rej_dep_{d_id}")
            )
            bot.send_message(call.message.chat.id, t, reply_markup=m)

    elif data.startswith("adm_appr_dep_"):
        dep_id = data.split("_")[3]
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT user_id, amount, status FROM deposits WHERE id=?", (dep_id,))
            row = cursor.fetchone()
            
            if row and row[2] == 'pending':
                u_id, amt, _ = row
                cursor.execute("UPDATE deposits SET status='approved' WHERE id=?", (dep_id,))
                cursor.execute("UPDATE users SET balance = balance + ? WHERE user_id=?", (amt, u_id))
                conn.commit()
                backup_db_to_telegram()
                
                bot.edit_message_text(f"✅ <b>Deposit Approved!</b> ({amt:.2f} BDT)", call.message.chat.id, call.message.message_id)
                try:
                    bot.send_message(u_id, f"🎉 আপনার ডিপোজিট এপ্রুভ হয়েছে এবং <b>{amt:.2f} BDT</b> অ্যাকাউন্টে যোগ হয়েছে!")
                except Exception:
                    pass
            
    elif data.startswith("adm_rej_dep_"):
        dep_id = data.split("_")[3]
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE deposits SET status='rejected' WHERE id=?", (dep_id,))
            conn.commit()
        backup_db_to_telegram()
        bot.edit_message_text("❌ <b>Deposit Rejected!</b>", call.message.chat.id, call.message.message_id)

    elif data == "adm_edit_payments":
        bk = get_setting('bkash_num')
        ng = get_setting('nagad_num')
        bn = get_setting('binance_uid')
        text = f"⚙️ <b>Edit Payment Methods</b>\n\n💸 bKash: {bk}\n💸 Nagad: {ng}\n💸 Binance UID: {bn}"
        markup = types.InlineKeyboardMarkup(row_width=1)
        markup.add(
            types.InlineKeyboardButton("💸 Edit bKash Number", callback_data="set_pay_bkash"),
            types.InlineKeyboardButton("💸 Edit Nagad Number", callback_data="set_pay_nagad"),
            types.InlineKeyboardButton("💸 Edit Binance UID", callback_data="set_pay_binance")
        )
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id, reply_markup=markup)

    elif data in ["set_pay_bkash", "set_pay_nagad", "set_pay_binance"]:
        m_name = data.split("_")[2]
        key = f"{m_name}_num" if m_name != "binance" else "binance_uid"
        set_user_state(user_id, "adm_step", f"set_setting:{key}")
        bot.send_message(call.message.chat.id, f"✏️ <b>নতুন {m_name.upper()} নম্বর / UID টি লিখে মেসেজ দিন:</b>")

    elif data == "adm_add_country":
        set_user_state(user_id, "adm_step", "add_country")
        bot.send_message(
            call.message.chat.id,
            "🌐 <b>নতুন সার্ভিস / কান্ট্রি (Range ID) যোগ করার ফরম্যাট:</b>\n\n"
            "<code>SERVICE_NAME,RANGE_ID,COUNTRY_NAME,FLAG,COUNTRY_CODE</code>\n\n"
            "উদাহরণ:\n"
            "<code>Facebook,26134,Uzbekistan,🇺🇿,uz</code>"
        )

    elif data == "adm_edit_balance":
        set_user_state(user_id, "adm_step", "edit_balance")
        bot.send_message(call.message.chat.id, "👤 <b>ইউজার ব্যালেন্স দিতে লিখুন:</b>\n\n<code>USER_ID,AMOUNT</code>")

    elif data == "adm_broadcast":
        set_user_state(user_id, "adm_step", "broadcast")
        bot.send_message(call.message.chat.id, "📢 <b>ইউজারদের কাছে পাঠাতে চাওয়া ব্রডকাস্ট মেসেজটি লিখুন:</b>")

# ----------------- GLOBAL TEXT & FILE HANDLER -----------------
@bot.message_handler(func=lambda m: True, content_types=['text', 'document'])
def global_message_handler(message):
    user_id = message.from_user.id
    txt = message.text.strip() if message.text else ""
    
    if user_id == ADMIN_ID:
        adm_step = get_user_state(user_id, "adm_step")
        if adm_step:
            if adm_step.startswith("set_setting:"):
                key = adm_step.split(":")[1]
                set_setting(key, txt)
                clear_user_state(user_id)
                bot.send_message(message.chat.id, f"✅ সফলভাবে <b>{key}</b> আপডেট হয়ে <b>{txt}</b> হয়েছে!")
                return

            elif adm_step == "add_country":
                try:
                    parts = txt.split(",")
                    srv_name = parts[0].strip()
                    srv_code = parts[1].strip()
                    c_name = parts[2].strip()
                    flag = parts[3].strip()
                    c_code = parts[4].strip()
                    
                    with get_db() as conn:
                        cursor = conn.cursor()
                        cursor.execute(
                            "INSERT INTO countries (service_name, service_code, country_name, country_flag, country_code, price) VALUES (?, ?, ?, ?, ?, 0.00)",
                            (srv_name, srv_code, c_name, flag, c_code)
                        )
                        conn.commit()
                    
                    backup_db_to_telegram()
                    clear_user_state(user_id)
                    bot.send_message(message.chat.id, f"✅ সফলভাবে <b>{flag} {c_name} ({srv_name})</b> সার্ভিস যোগ হয়েছে!")
                except Exception:
                    bot.send_message(message.chat.id, "❌ ফরম্যাট ভুল হয়েছে। ফরম্যাট: `SERVICE_NAME,RANGE_ID,COUNTRY_NAME,FLAG,COUNTRY_CODE`")
                return

            elif adm_step == "add_prod_info":
                try:
                    parts = txt.split(",")
                    cat = get_user_state(user_id, "adm_add_cat", "Mail")
                    
                    if cat == "VPN":
                        p_sub = parts[0].strip()
                        p_name = parts[1].strip()
                        p_price = float(parts[2].strip())
                    elif cat == "Proxy":
                        p_sub = parts[0].strip()
                        p_name = parts[1].strip()
                        p_price = float(parts[2].strip())
                    else:
                        p_sub = ""
                        p_name = parts[0].strip()
                        p_price = float(parts[1].strip())
                    
                    with get_db() as conn:
                        cursor = conn.cursor()
                        cursor.execute("SELECT id FROM products WHERE category=? AND name=? AND subcategory=?", (cat, p_name, p_sub))
                        res = cursor.fetchone()
                        
                        if res:
                            p_id = res[0]
                            cursor.execute("UPDATE products SET price=? WHERE id=?", (p_price, p_id))
                        else:
                            cursor.execute("INSERT INTO products (category, subcategory, name, price) VALUES (?, ?, ?, ?)", (cat, p_sub, p_name, p_price))
                            p_id = cursor.lastrowid
                        conn.commit()
                    
                    set_user_state(user_id, "active_pid", str(p_id))
                    set_user_state(user_id, "active_pname", p_name)
                    set_user_state(user_id, "adm_step", "await_stock_items")
                    
                    if cat == "Proxy":
                        instruct_stock = "🌐 <b>প্রক্সি আইটেমগুলোর তালিকা নিচে লিখুন বা ফাইল পাঠান:</b>\n\n<code>Address:Port:Username:Password</code>\n\nউদাহরণ:\n<code>192.168.1.1:8080:user123:pass123</code>"
                    else:
                        instruct_stock = f"✅ <b>{p_name} ({p_price:.2f} BDT)</b> সিলেক্ট হয়েছে!\n\nএখন একাউন্টগুলোর তালিকা পাঠিয়া স্টক দিন:"
                        
                    bot.send_message(message.chat.id, instruct_stock)
                except Exception:
                    bot.send_message(message.chat.id, "❌ ফরম্যাট ভুল হয়েছে। সঠিকভাবে আবার লিখুন।")
                return

            elif adm_step == "await_stock_items":
                p_id = get_user_state(user_id, "active_pid")
                p_name = get_user_state(user_id, "active_pname")
                lines = []
                
                if message.document:
                    file_info = bot.get_file(message.document.file_id)
                    downloaded_file = bot.download_file(file_info.file_path)
                    file_name = message.document.file_name or ""
                    
                    if file_name.endswith(('.xlsx', '.xls')):
                        try:
                            wb = openpyxl.load_workbook(io.BytesIO(downloaded_file))
                            ws = wb.active
                            for row in ws.iter_rows(values_only=True):
                                if row:
                                    row_vals = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                                    if row_vals:
                                        lines.append(" | ".join(row_vals))
                        except Exception as e:
                            bot.send_message(message.chat.id, f"❌ Excel ফাইল পড়তে সমস্যা: {e}")
                            return
                    else:
                        try:
                            text_data = downloaded_file.decode('utf-8', errors='ignore')
                        except Exception:
                            text_data = downloaded_file.decode('latin-1', errors='ignore')
                        lines = [line.strip() for line in text_data.split("\n") if line.strip()]
                else:
                    lines = [line.strip() for line in txt.split("\n") if line.strip()]

                if not lines:
                    bot.send_message(message.chat.id, "⚠️ কোনো অ্যাকাউন্ট/প্রক্সি পাওয়া যায়নি!")
                    return
                
                with get_db() as conn:
                    cursor = conn.cursor()
                    added_count = 0
                    for line in lines:
                        cursor.execute("INSERT INTO item_stock (product_id, content) VALUES (?, ?)", (p_id, line))
                        added_count += 1
                    conn.commit()
                
                backup_db_to_telegram()
                clear_user_state(user_id)
                bot.send_message(message.chat.id, f"🎉 সফলভাবে <b>{added_count} টি {p_name}</b> স্টকে যুক্ত হয়েছে!")
                return

            elif adm_step == "edit_balance":
                try:
                    parts = txt.split(",")
                    target_id = int(parts[0].strip())
                    add_bal = float(parts[1].strip())
                    
                    with get_db() as conn:
                        cursor = conn.cursor()
                        cursor.execute("UPDATE users SET balance = balance + ? WHERE user_id=?", (add_bal, target_id))
                        conn.commit()
                        
                    backup_db_to_telegram()
                    clear_user_state(user_id)
                    bot.send_message(message.chat.id, f"✅ ইউজার <code>{target_id}</code> এর ব্যালেন্সে <b>{add_bal:.2f} BDT</b> যোগ করা হয়েছে!")
                except Exception:
                    bot.send_message(message.chat.id, "❌ ফরম্যাট ভুল হয়েছে। ফরম্যাট: `USER_ID,AMOUNT`")
                return

            elif adm_step == "broadcast":
                with get_db() as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT user_id FROM users")
                    users = cursor.fetchall()
                s, f = 0, 0
                for u in users:
                    try:
                        bot.send_message(u[0], txt)
                        s += 1
                    except Exception:
                        f += 1
                clear_user_state(user_id)
                bot.send_message(message.chat.id, f"✅ ব্রডকাস্ট সম্পন্ন!\n\nসফল: {s} জন\nব্যর্থ: {f} জন")
                return

    usr_step = get_user_state(user_id, "step")
    if usr_step == "await_w_acc":
        acc_num = txt
        method = get_user_state(user_id, "w_method")
        
        set_user_state(user_id, "w_acc_final", acc_num)
        set_user_state(user_id, "step", "await_w_amt")
        
        bot.send_message(message.chat.id, f"💸 <b>Enter withdrawal amount in BDT for {method}:</b>")
        return

    elif usr_step == "await_w_amt":
        try:
            amt = float(txt)
            min_w = float(get_setting('min_withdraw') or 50.0)
            
            with get_db() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT reward_balance FROM users WHERE user_id=?", (user_id,))
                rew_row = cursor.fetchone()
                rew_bal = rew_row[0] if rew_row else 0.0
                
                if amt < min_w:
                    bot.send_message(message.chat.id, f"❌ সর্বনিম্ন উইথড্র {min_w:.2f} BDT।")
                    return
                if rew_bal < amt:
                    bot.send_message(message.chat.id, f"❌ আপনার পর্যাপ্ত রিওয়ার্ড ব্যালেন্স নেই! (বর্তমান রিওয়ার্ড ব্যালেন্স: {rew_bal:.2f} BDT)")
                    return
                    
                method = get_user_state(user_id, "w_method")
                acc = get_user_state(user_id, "w_acc_final")
                
                cursor.execute("UPDATE users SET reward_balance = reward_balance - ? WHERE user_id=?", (amt, user_id))
                cursor.execute("INSERT INTO withdrawals (user_id, amount, method, account_number) VALUES (?, ?, ?, ?)", (user_id, amt, method, acc))
                w_id = cursor.lastrowid
                conn.commit()

            backup_db_to_telegram()
            clear_user_state(user_id)
            bot.send_message(message.chat.id, f"✅ <b>আপনার {amt:.2f} BDT উইথড্র রিকুয়েস্ট সফলভাবে সাবমিট হয়েছে!</b>\n\nএডমিন ভেরিফাই করে আপনার একাউন্টে টাকা পাঠিয়ে দিবে।")
            
            admin_text = (
                f"📤 <b>NEW WITHDRAWAL REQUEST!</b>\n"
                f"━━━━━━━━━━━━━━━━━━\n"
                f"👤 <b>User ID:</b> <code>{user_id}</code>\n"
                f"💳 <b>Method:</b> {method}\n"
                f"📱 <b>Account:</b> <code>{acc}</code>\n"
                f"💵 <b>Amount:</b> <b>{amt:.2f} BDT</b>"
            )
            markup = types.InlineKeyboardMarkup(row_width=2)
            markup.add(
                types.InlineKeyboardButton("✅ Approve", callback_data=f"adm_appr_w_{w_id}"),
                types.InlineKeyboardButton("❌ Reject & Refund", callback_data=f"adm_rej_w_{w_id}")
            )
            try:
                bot.send_message(ADMIN_ID, admin_text, reply_markup=markup)
            except Exception:
                pass
        except Exception:
            bot.send_message(message.chat.id, "❌ সঠিক সংখ্যা টাইপ করুন।")
        return

    elif usr_step == "await_qty":
        try:
            qty = int(txt)
            p_id = get_user_state(user_id, "buying_p_id")
            
            with get_db() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT name, price, subcategory FROM products WHERE id=?", (p_id,))
                prod = cursor.fetchone()
                p_name, p_price, p_sub = prod[0], prod[1], prod[2]
                
                cursor.execute("SELECT COUNT(*) FROM item_stock WHERE product_id=? AND status='available'", (p_id,))
                p_stock = cursor.fetchone()[0]
                
                if qty <= 0 or qty > p_stock:
                    bot.send_message(message.chat.id, f"❌ পর্যাপ্ত স্টক নেই! সর্বোচ্চ {p_stock} টি নিতে পারবেন।")
                    return
                    
                total_bdt = qty * p_price
                total_usdt = total_bdt / 125.0
                
                cursor.execute("SELECT balance FROM users WHERE user_id=?", (user_id,))
                bal = cursor.fetchone()[0]
            
            clear_user_state(user_id)
            
            sub_info = f" ({p_sub})" if p_sub else ""
            text = (
                f"📬 <b>Order Summary</b>\n\n"
                f"⚡ <b>Item     :</b> {p_name}{sub_info}\n"
                f"👥 <b>Quantity :</b> {qty}\n"
                f"💰 <b>Total    :</b> {total_bdt:.2f} BDT / {total_usdt:.4f} USDT\n"
                f"💳 <b>Balance  :</b> {bal:.2f} BDT"
            )
            markup = types.InlineKeyboardMarkup(row_width=2)
            markup.add(
                types.InlineKeyboardButton("✅ Confirm Order", callback_data=f"cfmbuy_{p_id}_{qty}"),
                types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_action")
            )
            bot.send_message(message.chat.id, text, reply_markup=markup)
        except Exception:
            bot.send_message(message.chat.id, "❌ সঠিক সংখ্যা টাইপ করুন।")
        return

    elif usr_step == "await_dep_amt":
        try:
            amt = float(txt)
            min_dep = float(get_setting('min_deposit') or 20.0)
            if amt < min_dep:
                bot.send_message(message.chat.id, f"❌ সর্বনিম্ন ডিপোজিট {min_dep} BDT।")
                return
            
            method = get_user_state(user_id, "dep_method")
            num = get_user_state(user_id, "dep_num")
            clear_user_state(user_id)
            
            badge = "💸 <b>bKash Personal Number</b>" if method == "bkash" else ("💸 <b>Nagad Personal Number</b>" if method == "nagad" else "💸 <b>Binance Pay / UID</b>")
            
            text = (
                f"{badge}\n\n"
                f"Send <b>{amt:.2f} BDT</b> to:\n"
                f"<code>{num}</code>\n\n"
                f"<i>(নম্বরটি কপি করতে লেখার ওপর ট্যাপ করুন)</i>\n\n"
                f"টাকা পাঠানোর পর নিচের Paid বাটনে ক্লিক করুন:"
            )
            markup = types.InlineKeyboardMarkup(row_width=1)
            markup.add(
                types.InlineKeyboardButton("✅ Paid", callback_data=f"paid_{method}_{amt}"),
                types.InlineKeyboardButton("❌ Cancel", callback_data="cancel_action")
            )
            bot.send_message(message.chat.id, text, reply_markup=markup)
        except Exception:
            bot.send_message(message.chat.id, "❌ সঠিক সংখ্যা টাইপ করুন।")
        return

    elif usr_step == "await_trxid":
        trx_id = txt.upper()
        method = get_user_state(user_id, "dep_final_m")
        amt = float(get_user_state(user_id, "dep_final_a") or 20.0)
        
        dep_bonus_pct = float(get_setting('deposit_bonus') or 5.0)
        final_amt = amt + (amt * (dep_bonus_pct / 100.0))
        
        try:
            with get_db() as conn:
                cursor = conn.cursor()
                cursor.execute("INSERT INTO deposits (user_id, amount, method, trx_id) VALUES (?, ?, ?, ?)", (user_id, final_amt, method.upper(), trx_id))
                conn.commit()
                dep_id = cursor.lastrowid
            
            backup_db_to_telegram()
            clear_user_state(user_id)
            bot.send_message(message.chat.id, "✅ আপনার ডিপোজিট রিকুয়েস্ট সফলভাবে সাবমিট হয়েছে। এডমিন ভেরিফাই করে এপ্রুভ করে দেবে।")
            
            admin_text = (
                f"📥 <b>NEW DEPOSIT REQUEST!</b>\n"
                f"━━━━━━━━━━━━━━━━━━\n"
                f"👤 <b>User ID:</b> <code>{user_id}</code>\n"
                f"💳 <b>Method:</b> 💸 {method.upper()}\n"
                f"💵 <b>Amount:</b> {amt} BDT (+{dep_bonus_pct}% Bonus = <b>{final_amt:.2f} BDT</b>)\n"
                f"🏷️ <b>TrxID:</b> <code>{trx_id}</code>"
            )
            markup = types.InlineKeyboardMarkup(row_width=2)
            markup.add(
                types.InlineKeyboardButton("✅ Approve", callback_data=f"adm_appr_dep_{dep_id}"),
                types.InlineKeyboardButton("❌ Reject", callback_data=f"adm_rej_dep_{dep_id}")
            )
            try:
                bot.send_message(ADMIN_ID, admin_text, reply_markup=markup)
            except Exception:
                pass
        except sqlite3.IntegrityError:
            bot.send_message(message.chat.id, "❌ এই TrxID টি আগেই ব্যবহার করা হয়েছে!")
        return

# ----------------- 🔄 24/7 AUTO-RECONNECT ENGINE -----------------
print(f"🤖 {BOT_NAME} Started Successfully & Running 24/7...")
while True:
    try:
        bot.infinity_polling(timeout=20, long_polling_timeout=10)
    except Exception as e:
        print(f"⚠️ Connection dropped: {e}. Reconnecting in 5 seconds...")
        time.sleep(5)
